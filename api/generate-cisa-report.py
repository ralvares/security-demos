#!/usr/bin/env python3

import os
import requests
import json
import urllib3
import argparse
from tabulate import tabulate
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment

# Disable SSL certificate verification warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Check if the required environment variables are set
rox_api_token = os.environ.get('ROX_API_TOKEN')
rox_endpoint = os.environ.get('ROX_ENDPOINT')

if not rox_api_token or not rox_endpoint:
    print("Please set the 'ROX_API_TOKEN' and 'ROX_ENDPOINT' environment variables.")
    raise SystemExit

# Parse command line arguments
parser = argparse.ArgumentParser()
parser.add_argument('--deployment', help='Specify the deployment name')
parser.add_argument('--all', action='store_true', help='Scan all deployments')
parser.add_argument('--spreadsheet', action='store_true', help='Create spreadsheet instead of printing table')
args = parser.parse_args()

# Check if no arguments were specified
if not any(vars(args).values()):
    parser.print_help()
    raise SystemExit

known_exploited_vulnerabilities_url = 'https://www.cisa.gov/sites/default/files/feeds/known_exploited_vulnerabilities.json'
epss_api_url = 'https://api.first.org/data/v1/epss'

# Download known_exploited_vulnerabilities.json
response_known = requests.get(known_exploited_vulnerabilities_url)
catalog_known = response_known.json()

# Function to get the deployment ID by name
def get_deployment_id(deployment_name):
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {rox_api_token}'
    }
    params = {
        'query': f'Deployment:{deployment_name}'
    }
    url = f'https://{rox_endpoint}/v1/deployments'
    response = requests.get(url, headers=headers, params=params, verify=False)
    data = response.json()
    if 'deployments' in data and len(data['deployments']) > 0:
        return data['deployments'][0]['id']
    else:
        return None

# Function to get image information by deployment ID
def get_image_info(deployment_id):
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {rox_api_token}'
    }
    url = f'https://{rox_endpoint}/v1/deployments/{deployment_id}'
    response = requests.get(url, headers=headers, verify=False)
    return response.json()

# Function to download image information
def get_image(image_sha):
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {rox_api_token}'
    }

    url = f'https://{rox_endpoint}/v1/images/{image_sha}'
    response = requests.get(url, headers=headers, verify=False)
    return response.json()

# Function to get EPSS score and percentile for a CVE
def get_epss_data(cve):
    response_epss = requests.get(f'{epss_api_url}?cve={cve}')
    epss_data = response_epss.json()
    if epss_data['status'] == 'OK':
        return epss_data['data'][0]
    else:
        return None

# Function to scan a deployment
def scan_deployment(deployment_name):

    skipped_namespaces = ["openshift-", "stackrox", "rhacs"]
    if any(namespace in deployment_name for namespace in skipped_namespaces):
        #print(f"Skipping deployment '{deployment_name}' due to namespace.")
        return []
    
    # Get deployment ID by name
    deployment_id = get_deployment_id(deployment_name)
    if not deployment_id:
        print(f"Deployment '{deployment_name}' not found.")
        return []

    # Get image information by deployment ID
    image_info = get_image_info(deployment_id)
    if 'containers' not in image_info or len(image_info['containers']) == 0:
        print(f"No image information found for deployment '{deployment_name}'.")
        return []

    image_sha = image_info['containers'][0]['image']['id']

    # Get image information for a specific SHA256
    image_data = get_image(image_sha)

    # Extract image information
    image_id = image_data['id']
    image_full_name = image_data['name']['fullName']

    # Extract the list of vulnerabilities from known_exploited_vulnerabilities.json
    known_vulnerabilities = [vulnerability['cveID'] for vulnerability in catalog_known['vulnerabilities']]

    # Collect vulnerability information in a list of dictionaries
    vulnerabilities = []

    # Check if each vulnerability exists in the image information
    for component in image_data['scan']['components']:
        for vulnerability in component['vulns']:
            cve = vulnerability['cve']
            if cve in known_vulnerabilities:
                row = {
                    'Deployment Name': deployment_name,
                    'Image Full Name': image_full_name,
                    'CVE found on CISA Catalog': cve,
                    'Affected Component': component['name'],
                    'Affected Version': component['version'],
                    'Fixed By': component['fixedBy'] if component['fixedBy'] else 'No fix'
                }
                epss_data = get_epss_data(cve)
                if epss_data:
                    row['EPSS Score'] = epss_data['epss']
                    row['EPSS Percentile'] = epss_data['percentile']
                else:
                    row['EPSS Score'] = 'N/A'
                    row['EPSS Percentile'] = 'N/A'
                vulnerabilities.append(row)

    # Print the deployment and image information if CISA vulnerabilities are found
    if len(vulnerabilities) > 0:
        print(f"Generating report: Deployment={deployment_name} Image={image_full_name}")

    vulnerabilities.sort(key=lambda x: float(x['EPSS Score']) if x['EPSS Score'] != 'N/A' else 0, reverse=True)

    return vulnerabilities

# List to store vulnerability information for all deployments
all_vulnerabilities = []


def scan_multiple_deployments(deployments):
    vulnerabilities = []
    for deployment in deployments:
        skipped_namespaces = ["openshift-", "stackrox", "rhacs"]
        if any(skipped_namespace in deployment for skipped_namespace in skipped_namespaces):
            #print(f"Skipping deployment '{deployment}' due to namespace.")
            continue
        deployment_vulnerabilities = scan_deployment(deployment)
        vulnerabilities.extend(deployment_vulnerabilities)

    vulnerabilities.sort(key=lambda x: float(x['EPSS Score']) if x['EPSS Score'] != 'N/A' else 0, reverse=True)

    return vulnerabilities

# Scan deployments
if args.all:
    # Scan all deployments
    rox_api_token = os.environ.get('ROX_API_TOKEN')
    rox_endpoint = os.environ.get('ROX_ENDPOINT')

    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {rox_api_token}'
    }

    url = f'https://{rox_endpoint}/v1/deployments'
    response = requests.get(url, headers=headers, verify=False)
    data = response.json()
    if 'deployments' in data and len(data['deployments']) > 0:
        deployments = [deployment['name'] for deployment in data['deployments']]
        all_vulnerabilities = scan_multiple_deployments(deployments)
    else:
        print("No deployments found.")
else:
    # Scan specific deployment(s)
    deployments = args.deployment.split(',')
    all_vulnerabilities = scan_multiple_deployments(deployments)

# Create a spreadsheet with all the collected vulnerability information
if args.spreadsheet:
    df = pd.DataFrame(all_vulnerabilities)
    # Sort DataFrame by EPSS score in descending order
    df.sort_values(by='EPSS Score', ascending=False, inplace=True)

    risk_scores = []
    for row in all_vulnerabilities:
        score = row['EPSS Score']
        if score == 'N/A':
            risk_scores.append(0)
        else:
            score = float(score)
            if 0 <= score < 0.2:
                risk_scores.append('Low')
            elif 0.2 <= score < 0.4:
                risk_scores.append('Medium')
            elif 0.4 <= score < 0.6:
                risk_scores.append('High')
            else:
                risk_scores.append('Critical')
    df['Risk'] = risk_scores

    if args.deployment:
        spreadsheet_name = f'{args.deployment}_vulnerabilities.xlsx'
    else:
        spreadsheet_name = 'vulnerabilities.xlsx'
    
        # Create a new workbook
    workbook = Workbook()
    worksheet = workbook.active

    # Extract headers from DataFrame
    headers = df.columns.tolist()

    # Apply styling to the title row
    title_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
    title_alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
    title_row = [header.upper() for header in headers]
    worksheet.append(title_row)
    title_row_cells = worksheet[1]
    for cell in title_row_cells:
        cell.fill = title_fill
        cell.alignment = title_alignment

    # Apply styling to the data rows
    data_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    data_alignment = Alignment(vertical='center')

    for _, row in df.iterrows():
        row_values = row.tolist()
        worksheet.append(row_values)
        data_row_cells = worksheet[worksheet.max_row]
        for cell in data_row_cells:
            cell.fill = data_fill
            cell.alignment = data_alignment

    # Apply styling to the 'Risk' column
    risk_fill_low = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
    risk_fill_medium = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    risk_fill_high = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    risk_fill_critical = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    risk_column = worksheet['I']
    for cell in risk_column[1:]:
        value = cell.value
        if value == 'Low':
            cell.fill = risk_fill_low
        elif value == 'Medium':
            cell.fill = risk_fill_medium
        elif value == 'High':
            cell.fill = risk_fill_high
        elif value == 'Critical':
            cell.fill = risk_fill_critical

    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook as the final spreadsheet
    workbook.save(spreadsheet_name)
    print(f'Spreadsheet "{spreadsheet_name}" created.')
else:
    # Print the table for all deployments
    headers = all_vulnerabilities[0].keys()
    rows = [list(d.values()) for d in all_vulnerabilities]
    #print(tabulate(rows, headers=headers))
 
    # Add risk matrix based on EPSS score
    risk_scores = []
    for row in all_vulnerabilities:
        score = row['EPSS Score']
        if score == 'N/A':
            risk_scores.append(0)
        else:
            score = float(score)
            if 0 <= score < 0.2:
                risk_scores.append('Low')
            elif 0.2 <= score < 0.4:
                risk_scores.append('Medium')
            elif 0.4 <= score < 0.6:
                risk_scores.append('High')
            else:
                risk_scores.append('Critical')

    headers = list(all_vulnerabilities[0].keys()) + ['Risk']
    for i, row in enumerate(rows):
        row.append(risk_scores[i])

    print(tabulate(rows, headers=headers))